import nvdlib
import pandas as pd
from datetime import datetime
import time
from openpyxl import load_workbook
from openpyxl.styles import Font

# Function to search for a CVE using nvdlib
def search_cve(cve_number, api_key):
    cves = nvdlib.searchCVE(cveId=cve_number, key=api_key)
    if cves:
        for cve in cves:
            cve_id = cve.id
            description = cve.descriptions[0].value if cve.descriptions else 'N/A'

            # Handle both date formats, with or without fractional seconds
            try:
                published_date = datetime.strptime(cve.published, "%Y-%m-%dT%H:%M:%S.%f")
            except ValueError:
                published_date = datetime.strptime(cve.published, "%Y-%m-%dT%H:%M:%S")

            cve_age = (datetime.now() - published_date).days
            
            severity = cve.v31score if hasattr(cve, 'v31score') else 'N/A'

            return {
                "CVE Number": cve_id,
                "CVE Name": description,
                "CVE Age (Days)": cve_age,
                "CVE Severity": severity
            }
    else:
        print(f"No data found for CVE: {cve_number}")
        return None

# Function to read CVE numbers from a text file and skip the header
def read_cve_file(file_path):
    with open(file_path, 'r') as file:
        lines = file.readlines()
        return [line.strip() for line in lines[1:]]  # Skip the first line (header)

# Function to apply Calibri font to all cells in the Excel file
def apply_calibri_font(filename):
    wb = load_workbook(filename)
    ws = wb.active
    calibri_font = Font(name='Calibri')

    for row in ws.iter_rows():
        for cell in row:
            cell.font = calibri_font

    wb.save(filename)

# Function to save data to Excel and apply Calibri font
def save_to_excel(data, filename="output/cve_data.xlsx"):
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False)
    apply_calibri_font(filename)
    print(f"Data successfully saved to {filename}")

# Main Function with rate limiting
def main():
    api_key = '8da76f9f-b219-4b4a-80eb-0f22f07836a6'  # Replace with your actual API key
    cve_file = input("Enter the path to the CVE file (First line is expected to be a header then one CVE per line): ")
    cve_numbers = read_cve_file(cve_file)
    all_cve_data = []

    for cve_number in cve_numbers:
        print(f"Querying NVD: {cve_number}")
        cve_data = search_cve(cve_number, api_key)
        if cve_data:
            all_cve_data.append(cve_data)
        time.sleep(1)  # Rate limiting: 1 second between requests

    if all_cve_data:
        save_to_excel(all_cve_data)
    else:
        print("No CVE data found.")

if __name__ == "__main__":
    main()
